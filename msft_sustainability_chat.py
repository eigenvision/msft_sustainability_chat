
# IMPORTS
import streamlit as st
import openai

from llama_index.core import Settings

from llama_index.core import VectorStoreIndex
from llama_index.llms.openai import OpenAI
# from llama_index.llms import AzureOpenAI
from llama_index.core import SimpleDirectoryReader
from llama_index.core.llms import ChatMessage
# from llama_index.embeddings import AzureOpenAIEmbedding
from llama_index.core import StorageContext, load_index_from_storage
# from llama_index.core.postprocessor import SimilarityPostprocessor, LLMRerank

# OTHER IMPORTS
import os
import time
from datetime import datetime
import pytz
# import docx2txt
import pandas as pd
from openpyxl import load_workbook

# Q&A LOGGING FUNCTION

# In case we want to just log to a local csv file
# to keep log local use --local flag at the command line when launching script
local_mode = os.getenv('LOCAL_MODE') == 'True' # use "set LOCAL_MODE=True" to check if env variable set at command line

def log_chat_local(user_question, assistant_answer, response_time, model, words, prompt_type):
    file_path = r'C:\Users\Jeff\Dropbox\Jeff DropBox\Cloud Assurance\LLM Projects\Streamlit experiments\Streamlit chat logs\chat_log.xlsx'
    timezone = pytz.timezone("America/Los_Angeles")  # Pacific Time Zone
    current_time = datetime.now(timezone).strftime("%Y-%m-%d %H:%M:%S")
    
    # Data to log as a list of values
    data_to_log = [
        current_time, 
        user_question, 
        assistant_answer,
        response_time,
        model,
        words,
        prompt_type
        # str(chunk_params),  # Ensure complex objects are converted to string
        # str(query_params),   # Ensure complex objects are converted to string
        # search_time
    ]

    # Check if the file exists and is not empty
    if os.path.isfile(file_path):
        # Load the existing workbook and get the active sheet
        book = load_workbook(file_path)
        sheet = book.active

        # Find the first empty row in the first column
        for row in range(1, sheet.max_row + 2):
            if sheet.cell(row=row, column=1).value is None:
                break

        # Append the data row to the sheet
        for col, entry in enumerate(data_to_log, start=1):
            sheet.cell(row=row, column=col, value=entry)

        # Save the changes to the workbook
        book.save(file_path)
        book.close()

    else:
        # If the file doesn't exist, create it and add headers and the first row of data
        from openpyxl import Workbook
        book = Workbook()
        sheet = book.active
        headers = ['Date', 'User Question', 'Chatbot Answer', 'Response Time', 'Model', 'Word Count', 'System Prompt']
        sheet.append(headers)  # Append the headers
        sheet.append(data_to_log)  # Append the first row of data
        book.save(file_path)
        book.close()


# WORD COUNT FUNCTION
def count_words(text):
    # Remove leading and trailing whitespace
    text = text.strip()

    # Count the number of words in the text
    word_count = len(text.split())

    return word_count


# DEFINE THE SYSTEM PROMPTS

depth_prompt = """
        You are a helpful expert with deep technical and policy knowledge about the full range of Microsoft's 
        environmental sustainability policies and programs, especially those that concern Microsoft's
        commitment to become carbon negative, water positive, zero waste, and to protect more land than
        it uses.
        Your job is to provide maximally complete, detailed, logical, clear, and accurate answers to questions 
        about:
        1. Microsoft's environmental sustainability commitments and the policies and programs Microsoft is
        using to fulfill those commitments.
        2. The technologies and economic techniques Microsoft is applying or developing in pursuit of those commitments.
        3. The government and corporate policies that Microsoft recommends that governments and corporations
        should adopt in order to help the world achieve zero carbon and a truly sustainabile environment.
        You always pay close attention to the exact phrasing of the user's question and you always 
        deliver an answer that matches every specific detail of the user's expressed intention.
        You always give the fullest and most complete answers possible in substantial paragraphs 
        full of relevant details, concrete examples, and logical analysis.
        You never make up facts. You are not lazy and you do not skimp on important context.
        You always write in very polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        The users of this chatbot are making critical business decisions based on 
        your answers, and they need the highest quality answers possible that comply to the letter
        with their precise instructions.
        Finally, we have observed in your recent answers a distinct tendency to be lazy by providing
        answers that are short, lacking in extensive detail, and lacking in substantive analysis. It is 
        imperative that in this answer you overcome your instinct toward laziness and that you provide
        the richest, most detailed, most extensive answer that is possible to assemble from your
        sources, while still scrupulous respecting user instructions regarding the length of your response.
        YOU NEVER GIVE SHORT ANSWERS UNLESS SPECIFICALLY INSTRUCTED TO DO SO. In most cases a one
        paragraph answer is TOO SHORT.
        """

talking_points_prompt = """
        You are a helpful expert with deep technical and policy knowledge about the full range of Microsoft's 
        environmental sustainability policies and programs, especially those that concern Microsoft's
        commitment to become carbon negative, water positive, zero waste, and to protect more land than
        it uses.
        Your job is to provide substantive, detailed, accurate, and persuasive talking points about
        the goals, concrete details, and expected benefits of Microsoft's environmental sustainability policies and programs. 
        The audiences for these talking points are government policymakers, enterprise customers, tech industry partners, 
        non-profit organizations, the media, and other stakeholder. These talking points will typically cover
        some or all of the following areas, depending on the exact phrasing of the user's request:
        1. Microsoft's environmental sustainability commitments and the policies and programs Microsoft is
        using to fulfill those commitments.
        2. The technologies and economic techniques Microsoft is applying or developing in pursuit of those commitments.
        3. The government and corporate policies that Microsoft recommends that governments and corporations
        should adopt in order to help the world achieve zero carbon and a truly sustainabile environment.
        A set of talking points that you produce in response to a user request will typically consist of five to
        ten talking points, where each talking point consists of a succinct title and a substantive paragraph
        of finished text that explains in detail the main idea conveyed by the title.
        You always pay close attention to the exact phrasing of the user's question to determine the correct
        number of talking points for you to produce and their desired style and format.
        You always listen carefully for indications of the precise audience that the user wishes to
        address with the talking points you create.
        Your talking points make the most persuasive points possible in a clear logiccal order and include 
        relevant details, concrete examples, and logical analysis.
        You never make up facts. You are not lazy and you do not skimp on important context.
        You always write in very polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        The users of this chatbot are conducting critical business discussion based on 
        the talking points you create, and they need the highest quality answers possible that 
        comply strictly with their expressed or implied intentions regard content, style, and audience.
        Finally, we have observed in your recent answers a distinct tendency to be lazy by providing
        answers that are short, lacking in extensive detail, and lacking in substantive analysis. It is 
        imperative that in this answer you overcome your instinct toward laziness and that you provide
        the richest, most detailed, most extensive set of talking points that is possible to assemble from your
        sources, while still scrupulous respecting user instructions.
        """

slides_prompt = """
        You are a helpful expert with deep technical and policy knowledge about the full range of Microsoft's 
        environmental sustainability policies and programs, especially those that concern Microsoft's
        commitment to become carbon negative, water positive, zero waste, and to protect more land than
        it uses.
        Your job is to create the complete text content for slide deck presentations about
        the goals, concrete details, and expected benefits of Microsoft's environmental sustainability policies and programs. 
        The audiences for these slide decks are government policymakers, enterprise customers, tech industry partners, 
        non-profit organizations, the media, and other stakeholder. 
        Your responses will always contain the full finished text content for a 10 slide presentation unless the user
        requests a different number of slides. The content you create for each slide must always exactly follow
        this three part format:
        SLIDE TITLE: a concise and relevant title that states the main idea of the slide in clear and easy to understand fashion.
        SLIDE BODY: a minimum of three and a maximum of five concise bullet points that detail specific parts of the slide's main idea.
        SLIDE SPEAKER NOTES: two or preferably more detailed, substantive, extensive, multi-sentence paragraphs providing a step by step
        explanation of each of the bullet points on the slide and emphasize the slide's main idea.
        The set of slides in a complete presentation will always fit together into a coherent whole where the reason for the 
        presence of each slide is obvious to the audience.
        The slides you create will typically cover some or all of the following areas, depending on the exact phrasing of the user's request:
        1. Microsoft's environmental sustainability commitments and the policies and programs Microsoft is
        using to fulfill those commitments.
        2. The technologies and economic techniques Microsoft is applying or developing in pursuit of those commitments.
        3. The government and corporate policies that Microsoft recommends that governments and corporations
        should adopt in order to help the world achieve zero carbon and a truly sustainabile environment.
        You always listen carefully for indications of the precise audience that the user wishes to
        address with the slides you create.
        You never make up facts. You are not lazy and you do not skimp on important context.
        You always write in very polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        The users of this chatbot are conducting critical business discussion based on 
        the talking points you create, and they need the highest quality presentations possible that 
        comply strictly with their expressed or implied intentions regard content, style, and audience.
        Finally, we have observed in your recent answers a distinct tendency to be lazy by providing
        responses that are short, lacking in extensive detail, and lacking in substantive analysis. It is 
        imperative that in this answer you overcome your instinct toward laziness and that you provide
        the richest, most detailed, most extensive set of slides that is possible to assemble from your
        sources, while still scrupulous respecting user instructions.
        """

quick_prompt = """
        You are a helpful expert with deep technical and policy knowledge about the full range of Microsoft's 
        environmental sustainability policies and programs, especially those that concern Microsoft's
        commitment to become carbon negative, water positive, zero waste, and to protect more land than
        it uses. 
        Your job is to provide concise and accurate answers to questions about Microsoft's environmental
        sustainability policies and programs.
        You pay close attention to the phrasing of the user's question.
        You never make up facts.
        You always provide answers that are quick and too the point, without unecessary
        explanations or words.
        You always write in polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        DO NOT explicitly mention the conversation you are engaged in. Just aanswer the user's question.
        """


# SET UP STREAMLIT
st.set_page_config(page_title="Chat with the 2024 Microsoft Environmental Sustainability Report", layout="centered", initial_sidebar_state="auto", menu_items=None)

# Inject custom CSS to center the title using the correct class name
# Inject custom CSS for styling the title
st.markdown("""
<style>
.custom-title {
    font-size: 48px;  /* Approximate size of Streamlit's default title */
    font-weight: bold; /* Streamlit's default title weight */
    line-height: 1.25; /* Reduced line height for tighter spacing */
    color: rgb(0, 163, 238); /* Default color, adjust as needed */
    text-align: center;
    margin-bottom: 30px; /* Adjust the space below the title */
}
</style>
""", unsafe_allow_html=True)

# Custom title with a line break using HTML in markdown
st.markdown('<div class="custom-title">Chat with 2024 Microsoft Environmental Sustainability Report</div>', unsafe_allow_html=True)

st.info("""This chatbot lets you converse with the 2024 Microsoft Environmental Sustainability Report.
    Using GPT-4, the chatbot is designed to produce substantive and accurate answers about Microsoft's
    sustainability work suitable for use in customer and stakeholder engagements. Nevertheless, because AI
    can sometimes produce unexpected or wrong answers, you should always review carefully before sharing. 
    If you're not satisfied with the AI's first attempt, try revising your prompt or asking follow up questions.
      \n\nIf you're asking for a complex response like a set of talking points or presentation slides, try to 
    give the bot an outline of the ideas or topics you want it to cover. You should also tell it how many
    talking points or slides you expect.""")

# Custom CSS to center text
st.markdown("""
<style>
.centered-text {
    text-align: center;
}
</style>
""", unsafe_allow_html=True)

# Text with hyperlink centered
st.markdown('<p class="centered-text">Download the <a href="https://query.prod.cms.rt.microsoft.com/cms/api/am/binary/RW1lhhu" target="_blank">2024 Microsoft Environmental Sustainability Report</a></p>', unsafe_allow_html=True)

# SELECT SYSTEM PROMPT AND MODEL
# Define the possible choices linking to the variables

# Define the mappings for system prompts and model versions
style_settings = {
    "Quick answers": {
        "s_prompt": quick_prompt,
        "model": "gpt-3.5-turbo"
    },
    "In-depth answers": {
        "s_prompt": depth_prompt,
        "model": "gpt-4-turbo"
    },
    "Talking points": {
        "s_prompt": talking_points_prompt,
        "model": "gpt-4-turbo"
    },
    "Presentation slides (text)": {
        "s_prompt": slides_prompt,
        "model": "gpt-4-turbo"
    }
}

# Use the sidebar to create radio buttons for system_prompt and model selection
selected_style = st.sidebar.radio("Choose what kind of answer you want:", list(style_settings.keys()))

# Get the selected system prompt and model
system_prompt = style_settings[selected_style]["s_prompt"]
model = style_settings[selected_style]["model"]
# enable logging of the selected radio button
prompt_type = selected_style

# SET UP LLM AND PARAMETERS
openai.api_key = st.secrets.openai_key
# "gpt-4-turbo" # "gpt-4-0125-preview" "gpt-4-1106-preview" "gpt-4-turbo" "gpt-3.5-turbo"
# models: gpt-4-turbo pt-4-1106-preview, gpt-4 (gpt-4-0613), gpt-4-32k (gpt-4-32k-0613) but "model not found", 
# gpt-3.5-turbo, gpt-3.5-turbo-instruct, gpt-35-turbo-16k

llm=OpenAI(
        api_key = openai.api_key,
        model=model,
        temperature=0.5,
        system_prompt=system_prompt)


# SET UP LLAMAINDEX SETTINGS
# service_context deprecated
chunk_size=512
chunk_overlap=256
chunk_params = (chunk_size, chunk_overlap)
Settings.llm = llm
Settings.chunk_size = chunk_size
Settings.chunk_overlap = chunk_overlap

if "messages" not in st.session_state.keys(): # Initialize the chat messages history
    st.session_state.messages = [
        {"role": "assistant", "content": """Ask a question and use the buttons on the left to 
         tell the chatbot what kind of answer you want: 
         """}
    ]


# LOAD DOCS INTO LLAMAINDEX, CREATE INDEX (AND RELOAD IF IT ALREADY EXISTS)
@st.cache_resource(show_spinner=False)
def load_data():
  with st.spinner(text="Loading the reports. This will take a few minutes."):
      # Define the path to the index file
      persist_dir = './index'

      # Check if the index directory exists, create if not
      if not os.path.exists(persist_dir):
            os.makedirs(persist_dir)

        # Now attempt to load or create the index
      try:
            storage_context = StorageContext.from_defaults(persist_dir=persist_dir)
            index = load_index_from_storage(storage_context)
      except FileNotFoundError:
            # Index not found, create it
            reader = SimpleDirectoryReader(input_dir="./msft_sustainability_docs")
            docs = reader.load_data()
            index = VectorStoreIndex.from_documents(docs)

            # Save the index to the file
            index.storage_context.persist(persist_dir=persist_dir)

      return index

index = load_data()

# DEFINE THE RUN_CHATS FUNCTION
def run_chats(query):
    
    search_time = 0.0 

    similarity_top_k = 15
   
    start_time_search = time.time() # time vector search
    chat_engine = index.as_chat_engine(chat_mode="condense_question",
                                       streaming=True,
                                       similarity_top_k=similarity_top_k, 
                                   ) # verbose=True # streaming=True
    end_time_search = time.time()
           
    result = chat_engine.chat(query)  # chat_engine.stream_chat(query)
    # Calculate search time
    search_time = end_time_search - start_time_search

    # Store the values of k
    query_params = similarity_top_k
    # result.print_response_stream()
    
    return result, query_params, search_time


# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
# Create the selected version of the chat engine

if "chat_engine" not in st.session_state.keys(): # Initialize the chat engine
        st.session_state.chat_engine = None

if prompt := st.chat_input("Your question"): # Prompt for user input and save to chat history
    st.session_state.messages.append({"role": "user", "content": prompt})

for message in st.session_state.messages: # Display the prior chat messages
    with st.chat_message(message["role"]):
        st.write(message["content"])

# If last message is not from assistant, generate a new response
if st.session_state.messages[-1]["role"] != "assistant":
    with st.chat_message("assistant"):
        with st.spinner("Thinking..."):
            start_time = time.time()
            response, query_params, search_time = run_chats(st.session_state.messages[-1]["content"])
            response_time = time.time() - start_time
            st.write(response.response)
            message = {"role": "assistant", "content": response.response}
            st.session_state.messages.append(message) # Add response to message history
            words = count_words(response.response)
            prompt_type = str(prompt_type)
            log_chat_local(user_question=prompt,
                               assistant_answer=response.response,
                               response_time=response_time,
                               model=model,
                               words=words,
                               prompt_type=prompt_type
                            )



